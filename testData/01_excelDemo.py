import openpyxl

book = openpyxl.load_workbook("C:\\automation\excel\python_excel.xlsx")
lst = []
sheet = book.active # active means active sheet on application
cell = sheet.cell(row=1, column=2) # cell attraction
print(cell.value) # Read value from that mention particuler name

sheet.cell(row=2,column=2).value = "Abhi" # Writitng int the mentioned cell
print(sheet.cell(row=2,column=2).value)

print(sheet.max_row)
print(sheet.max_column)
print(sheet['C1'].value) # Anoter reading methode
print("***************")
for i in range(2, (sheet.max_row + 1)): # raws
    # if sheet.cell(row=i, column=1).value == "testcase2":
    Dict = {}
    for j in range(2, (sheet.max_column + 1)): # column
        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    lst.append(Dict)


print(lst)