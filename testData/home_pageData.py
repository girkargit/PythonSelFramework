import openpyxl


class HomePageData:
    # test_home_page_data = [{'firstname': 'abhi', 'email': 'test@test.com', 'gender': 'Male'}, {'firstname': 'suraj', 'email': 'test@test.com', 'gender': 'Male'}]
    @staticmethod
    def get_testData():
        lst = [] # Above format store in lst
        book = openpyxl.load_workbook("C:\\automation\excel\python_excel.xlsx")
        sheet = book.active  # active means active sheet on application
        for i in range(2, (sheet.max_row + 1)):  # raws
            Dict = {}
            for j in range(2, (sheet.max_column + 1)):  # column
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            lst.append(Dict)
        return lst